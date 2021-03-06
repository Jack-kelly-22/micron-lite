from openpyxl import Workbook,load_workbook
from math import pi
from PIL import Image
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import constants
from openpyxl.drawing.image import Image
class SpreadWriter:

    def __init__(self,filter_dic):
        self.job_name = filter_dic['job_name']
        print("start spread init")
        self.constants = filter_dic['options']['constants']
        # self.frame_name = frame.name
        self.filename = filter_dic['out_path']
        self.workbook = load_workbook("Output_Sheet_Blank.xlsx")
        self.i = 15
        self.page = self.workbook.active
        self.write_ref_settings(filter_dic)
        self.write_inspect_spec(filter_dic)
        self.write_inspect_results(filter_dic)
        self.write_rejected_images(filter_dic)

    def write_ref_settings(self,filter_dic):
        """writes contents of fixture reference settings"""
        #B6: software version
        #B7: scale
        #B8: #of images
        #B9: threshold method
        #B10: threshold value
        self.page['B6'] = constants.get_version()
        self.page['B7'] = constants.get_scale()
        self.page['B8'] = filter_dic['num_images']
        self.page['B9'] = str(self.constants['thresh'])
        self.page['B10'] = str(self.constants['thresh'])


    def write_inspect_spec(self,filter_dic):
        """writes inspection specifications"""
        #D6 max pore size(um)
        #E7 Minimum porosity(%)
        self.page['D6'] = filter_dic['max_pore_size']['diam']
        self.page['E7'] = str(filter_dic['max_porosity']) + '%'

    def write_inspect_results(self,filter_dic):
        """Writes result data for each image"""
        #(i)A: Image Name
        #(i)B: Observed Porosity
        #(i)C: Max Observed Pore Size
        #(i)D: Image Result(P/F)
        for img_dic in filter_dic['images']:
            self.page['A' + str(self.i)] = img_dic['img_name']
            self.page['B' + str(self.i)] = img_dic['porosity']
            self.page['C' + str(self.i)] = img_dic['largest_pore'][1]
            # self.page[str(self.i)+'D']
            self.i=+1

    def rejected_header(self,image_dic):
        self.page['A' + str(self.i)] = "Image Ref: "
        self.page["B" + str(self.i)] = image_dic['img_name']
        self.i+=1
        self.page['A' + str(self.i)] = "Rejected Pore Diameter(μm)"
        self.page[+'B' + str(self.i)] = "Location"
        self.i+=1



    def write_rejected_images(self,filter_dic):
        self.page[str(self.i)+'A'] = "Rejected Images- Oversized Pore Observations"
        for image in filter_dic['images']:
            if not image['pass']:
                self.rejected_header(image)
                #add rejected image into spreadsheet
                fail_img = Image('.' + image['out_path'])
                fail_img.anchor = 'C' + str(self.i)
                self.page.add_image(fail_img)
                j=0
                for pore in image['violated_pores']:
                    self.page['A'+str(self.i+j)] = pore[1]
                    self.page['A'+str(self.i+j)] = '( '+str(pore[0][0])+', '+ str(pore[0][1]) + " )"
                    j += 1
                if j>15:
                    self.i += j
                else:
                    self.i += 15


